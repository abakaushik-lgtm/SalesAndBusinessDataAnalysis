import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def build_excel_workbook():
    csv_path = os.path.join("dataset", "sales_data.csv")
    if not os.path.exists(csv_path):
        print(f"Error: {csv_path} not found.")
        return
        
    os.makedirs("excel", exist_ok=True)
    excel_path = os.path.join("excel", "sales_analysis_model.xlsx")
    
    df = pd.read_csv(csv_path)
    
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Styling definitions
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    title_font = Font(name=font_family, size=14, bold=True, color="1E293B")
    metric_title_font = Font(name=font_family, size=10, bold=False, color="64748B")
    metric_val_font = Font(name=font_family, size=16, bold=True, color="0F172A")
    kpi_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # ---------------------------------------------------------
    # Sheet 1: Executive KPI Dashboard (Excel Formulas)
    # ---------------------------------------------------------
    ws_kpi = wb.create_sheet(title="Executive_KPIs")
    ws_kpi.views.sheetView[0].showGridLines = True
    
    ws_kpi["B2"] = "Global Retail Corp - Executive Financial Summary"
    ws_kpi["B2"].font = title_font
    
    kpis = [
        ("Total Enterprise Revenue", "=SUM(Raw_Data!R2:R10001)", "$#,##0.00", "C4", "C5"),
        ("Total Net Profit", "=SUM(Raw_Data!S2:S10001)", "$#,##0.00", "E4", "E5"),
        ("Overall Profit Margin %", "=E5/C5", "0.00%", "G4", "G5"),
        ("Total Order Transactions", "=COUNTA(Raw_Data!A2:A10001)", "#,##0", "I4", "I5"),
        ("Average Order Value (AOV)", "=C5/I5", "$#,##0.00", "K4", "K5"),
        ("Average Discount Rate", "=AVERAGE(Raw_Data!Q2:Q10001)", "0.00%", "M4", "M5")
    ]
    
    for title, formula, num_fmt, label_cell, val_cell in kpis:
        ws_kpi[label_cell] = title
        ws_kpi[label_cell].font = metric_title_font
        ws_kpi[label_cell].alignment = Alignment(horizontal="center")
        
        ws_kpi[val_cell] = formula
        ws_kpi[val_cell].font = metric_val_font
        ws_kpi[val_cell].number_format = num_fmt
        ws_kpi[val_cell].alignment = Alignment(horizontal="center")
        
    # Category Pivot Table Simulation via Excel SUMIFS Formulas
    ws_kpi["B8"] = "Product Category Performance Breakdown (SUMIFS Modeling)"
    ws_kpi["B8"].font = Font(name=font_family, size=12, bold=True, color="1E293B")
    
    cat_headers = ["Category", "Total Units Sold", "Gross Sales", "Net Profit", "Profit Margin %", "Avg Discount"]
    for col_num, h in enumerate(cat_headers, start=2):
        cell = ws_kpi.cell(row=10, column=col_num, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    categories = ["Technology", "Furniture", "Office Supplies"]
    for idx, cat in enumerate(categories, start=11):
        ws_kpi.cell(row=idx, column=2, value=cat).alignment = Alignment(horizontal="left")
        
        # Units Sold SUMIFS
        cell_units = ws_kpi.cell(row=idx, column=3, value=f'=SUMIFS(Raw_Data!O2:O10001, Raw_Data!L2:L10001, "{cat}")')
        cell_units.number_format = "#,##0"
        
        # Sales SUMIFS
        cell_sales = ws_kpi.cell(row=idx, column=4, value=f'=SUMIFS(Raw_Data!R2:R10001, Raw_Data!L2:L10001, "{cat}")')
        cell_sales.number_format = "$#,##0.00"
        
        # Profit SUMIFS
        cell_profit = ws_kpi.cell(row=idx, column=5, value=f'=SUMIFS(Raw_Data!S2:S10001, Raw_Data!L2:L10001, "{cat}")')
        cell_profit.number_format = "$#,##0.00"
        
        # Margin %
        cell_margin = ws_kpi.cell(row=idx, column=6, value=f'=E{idx}/D{idx}')
        cell_margin.number_format = "0.00%"
        
        # Avg Discount AVERAGEIFS
        cell_disc = ws_kpi.cell(row=idx, column=7, value=f'=AVERAGEIFS(Raw_Data!Q2:Q10001, Raw_Data!L2:L10001, "{cat}")')
        cell_disc.number_format = "0.00%"

    # ---------------------------------------------------------
    # Sheet 2: Raw Data Sheet
    # ---------------------------------------------------------
    ws_raw = wb.create_sheet(title="Raw_Data")
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_raw.append(r)
        
    # Style Header Row of Raw Data
    for cell in ws_raw[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    # Auto-fit column widths
    for ws in [ws_kpi, ws_raw]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(excel_path)
    print(f"Successfully created Excel financial model at {excel_path}")

if __name__ == "__main__":
    build_excel_workbook()
